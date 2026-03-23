import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import re

def clean_text(text):
    if not text: return ""
    text = str(text).replace("<br>", " ").strip()
    return re.sub(r'\s+', ' ', text)

def find_value_in_json(header, data):
    """
    Looks for the header text within the extracted JSON cells.
    If found, tries to return either the cell to the right or the cell directly underneath.
    """
    if not header: return ""
    header_lower = str(header).lower().strip()
    pages = data.get("document", {}).get("pages", [])
    
    for page in pages:
        for element in page.get("elements", []):
            if element.get("type") == "table":
                rows = element.get("rows", [])
                for r_idx, row in enumerate(rows):
                    row_texts = [clean_text(cell.get("text", "")) if isinstance(cell, dict) else "" for cell in row]
                    for c_idx, text in enumerate(row_texts):
                        text_lower = text.lower()
                        if header_lower in text_lower:
                            # Direct string match inside the cell, meaning they might be split by colon (e.g. "Bill No: 123")
                            parts = text_lower.split(header_lower, 1)
                            if len(parts) > 1 and parts[1].strip() and parts[1].strip() not in [":", "-"]:
                                raw_extracted = text[len(parts[0]) + len(header_lower):].strip(" :-=\n")
                                if raw_extracted:
                                    return raw_extracted
                            
                            # Check cell to the right
                            if c_idx + 1 < len(row_texts) and row_texts[c_idx + 1].strip():
                                return row_texts[c_idx + 1].strip()
                            
                            # Check cell below
                            if r_idx + 1 < len(rows):
                                next_row_texts = [clean_text(cell.get("text", "")) if isinstance(cell, dict) else "" for cell in rows[r_idx+1]]
                                if c_idx < len(next_row_texts) and next_row_texts[c_idx].strip():
                                    return next_row_texts[c_idx].strip()
                            return ""
    return ""

def extract_bescom_to_excel(json_path, output_excel, template_path=None):
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"JSON file not found: {json_path}")
        
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Use template if provided
    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Read headers from row 1
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value:
                headers.append((col_idx, str(cell.value)))
                
        # Fill row 2
        for col_idx, header in headers:
            val = find_value_in_json(header, data)
            if val:
                out_cell = ws.cell(row=2, column=col_idx)
                out_cell.value = val
                out_cell.font = Font(name="Calibri", size=11)
                out_cell.alignment = Alignment(wrap_text=True)
                out_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    else:
        # Fallback to empty if no template provided
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "No Template Uploaded"
        
    wb.save(output_excel)
    return output_excel
