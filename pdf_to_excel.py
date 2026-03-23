"""
pdf_to_excel.py (V5.2 – Content-Aware Extraction)
================================================
Fixes "Empty Tabs" by performing a pre-write content density check.
Ensures only tables with meaningful text are converted to Excel sheets.
"""

from __future__ import annotations
import re
import os
import traceback
from typing import Any

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 – STYLING
# ─────────────────────────────────────────────────────────────────────────────

_S_THIN = Side(style="thin",   color="000000")
_S_MED  = Side(style="medium", color="000000")

_B_THIN  = Border(left=_S_THIN, right=_S_THIN, top=_S_THIN, bottom=_S_THIN)
_B_MED   = Border(left=_S_MED,  right=_S_MED,  top=_S_MED,  bottom=_S_MED)
_B_TITLE = Border(left=_S_THIN, right=_S_THIN, top=_S_THIN, bottom=_S_MED)

_F_TITLE  = Font(name="Calibri", bold=True,  size=12)
_F_HDR    = Font(name="Calibri", bold=True,  size=10)
_F_KEY    = Font(name="Calibri", bold=True,  size=10)
_F_NORM   = Font(name="Calibri", bold=False, size=10)

_A_WRAP_TL = Alignment(wrap_text=True, vertical="top",    horizontal="left")
_A_CTR     = Alignment(wrap_text=True, vertical="center", horizontal="center")
_A_CTR_L   = Alignment(wrap_text=True, vertical="center", horizontal="left")

def _set_col(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── Table Strategies ──────────────────────────────────────────────────────────
_STRATEGIES = [
    {"vertical_strategy": "lines", "horizontal_strategy": "lines", "snap_tolerance": 4},
    {"vertical_strategy": "lines", "horizontal_strategy": "text",  "snap_tolerance": 5},
    {"vertical_strategy": "text",  "horizontal_strategy": "lines", "snap_tolerance": 5},
]

_FIELD_GROUPS = {
    "Billing Info":     ["RR No", "Account ID", "Bill No", "Billing Period", "Bill Date", "Due Date", "Disconnection Date"],
    "Customer Info":    ["Name & Address", "Net Payable Amount"],
    "Technical Info":   ["Tariff", "Contract Demand(KVA)", "Billing Demand (KVA)", "Meter ID"],
}


def _norm(s: Any) -> str:
    if s is None: return ""
    return str(re.sub(r"\s+", " ", str(s))).strip()

def normalize_val(v: str) -> str:
    v = _norm(v)
    if not v: return v
    clean_n = v.replace(",", "").replace("Rs.", "").strip()
    if re.fullmatch(r"(\d+)\.?(\d*)", clean_n):
        try:
            parts = clean_n.split(".")
            integer_p = f"{int(parts[0]):,}"
            return f"{integer_p}.{parts[1]}" if len(parts)>1 and parts[1] else integer_p
        except: pass
    return v

def _iou(b1, b2) -> float:
    ix0, iy0 = max(b1[0], b2[0]), max(b1[1], b2[1])
    ix1, iy1 = min(b1[2], b2[2]), min(b1[3], b2[3])
    inter = max(0, ix1 - ix0) * max(0, iy1 - iy0)
    if inter <= 0: return 0.0
    a1 = (b1[2]-b1[0])*(b1[3]-b1[1])
    a2 = (b2[2]-b2[0])*(b2[3]-b2[1])
    return inter / (a1 + a2 - inter)

def _classify_name(data: list[list[str]], counter: int) -> str:
    flat = " ".join(c.lower() for r in data[:2] for c in r if c).strip()
    if "meter reading" in flat: return "Meter Readings"
    if "tod" in flat: return "TOD Readings"
    if "billing" in flat or "description" in flat: return "Billing Details"
    if "payment" in flat or "neft" in flat or "bank" in flat: return "Bank Details"
    return f"Table_{counter}"

def _has_content(data: list[list[str]], min_cells=3) -> bool:
    """Check if table has actual text data, not just empty boxes."""
    count = 0
    for row in data:
        for cell in row:
            if cell and len(str(cell).strip()) > 1:
                count = count + 1
            if count >= min_cells: return True
    return False

def extract_v5_2(pdf_path: str):
    summary, tables, audit = {}, [], []
    used_bboxes = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            pg_txt = page.extract_text() or ""
            pg_lns = [ln.strip() for ln in pg_txt.splitlines() if ln.strip()]
            audit.extend(pg_lns)
            
            # --- 1. Greedy Regex labels ---
            matches = re.findall(r"^(.{3,40}?)\s*[:\.]{1,}\s*(.{1,120})$", pg_txt, re.M)
            for k, v in matches:
                k_clean = _norm(k)
                if k_clean not in summary and 3 < len(k_clean) < 35:
                    summary[k_clean] = normalize_val(v)

            # --- 2. Tables with Content Check ---
            for strat in _STRATEGIES:
                for ft in page.find_tables(table_settings=strat):
                    if any(_iou(ft.bbox, prev) > 0.5 for prev in used_bboxes): continue
                    
                    rows = ft.extract() or []
                    cleaned = [[_norm(c) for c in r] for r in rows if any(str(c).strip() for c in r)]
                    
                    if _has_content(cleaned):
                        used_bboxes.append(ft.bbox)
                        tables.append({
                            "name": _classify_name(cleaned, len(tables)+1),
                            "data": cleaned,
                            "y": ft.bbox[1] + (page.page_number * 1000)
                        })

    tables.sort(key=lambda t: t["y"])
    return summary, tables, audit

def _write_table(ws, data):
    ws.sheet_view.showGridLines = False
    if not data: return
    # Write Header
    for ci, val in enumerate(data[0], start=1):
        c = ws.cell(row=1, column=ci, value=val)
        c.font, c.alignment, c.border = _F_HDR, _A_CTR, _B_TITLE
        _set_col(ws, ci, 22)
    # Write Body
    for ri, row in enumerate(data[1:], start=2):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font, c.alignment, c.border = _F_NORM, _A_WRAP_TL, _B_THIN

def _write_summary(ws, summary, pdf_name):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:B1")
    ws["A1"].value, ws["A1"].font, ws["A1"].alignment, ws["A1"].border = f"PDF Data Summary: {os.path.basename(pdf_name)}", _F_TITLE, _A_CTR, _B_MED
    
    r = 3
    for group, fields in _FIELD_GROUPS.items():
        present = [(f, summary[f]) for f in fields if f in summary]
        if not present: continue
        ws.merge_cells(f"A{r}:B{r}")
        ws.cell(row=r, column=1, value=group).font = Font(bold=True, italic=True)
        r += 1
        for k, v in present:
            ws.cell(row=r, column=1, value=k).font = _F_KEY
            ws.cell(row=r, column=1).border = _B_THIN
            ws.cell(row=r, column=2, value=v).font = _F_NORM
            ws.cell(row=r, column=2).border = _B_THIN
            r += 1
        r += 1
    _set_col(ws, 1, 35); _set_col(ws, 2, 60)

def convert_pdf_to_excel(pdf_path: str, output_path: str) -> None:
    try:
        summary, tables, audit = extract_v5_2(pdf_path)
        wb = Workbook()
        _write_summary(wb.active, summary, pdf_path)
        for tbl in tables:
            ws = wb.create_sheet(title=str(tbl["name"])[:30])
            _write_table(ws, tbl["data"])
        ws_audit = wb.create_sheet("Document Text Audit")
        for i, ln in enumerate(audit, start=1):
            ws_audit.cell(row=i, column=1, value=ln).font = _F_NORM
        _set_col(ws_audit, 1, 120)
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"[V5.2] Saved: {output_path} ({len(tables)} active tables)")
    except Exception as e:
        print(f"[V5.2] Error: {e}")
        raise e
if __name__ == "__main__":
    import sys
    import glob

    # Default logic: Use path if provided as argument, else find latest PDF in ../uploads/ or uploads/
    pdf_input = sys.argv[1] if len(sys.argv) > 1 else None
    
    if not pdf_input:
        # Search for PDFs in 'uploads' directory relative to script or current dir
        search_paths = ["uploads/*.pdf", "../uploads/*.pdf", "pdfhtmlcodex/pdf_html_converter/uploads/*.pdf"]
        found_files = []
        for p in search_paths:
            found_files.extend(glob.glob(p))
            
        if found_files:
            # Sort by modification time to get the latest
            found_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            pdf_input = str(found_files[0])
            print(f"[*] No input specified. Using latest PDF: {pdf_input}")
        else:
            print("[!] Error: No PDF files found in 'uploads/' and no file specified.")
            sys.exit(1)

    # Generate output path
    base_name = os.path.splitext(os.path.basename(pdf_input))[0]
    out_path = os.path.join("output", f"{base_name}.xlsx")
    
    if pdf_input:
        print(f"[*] Converting: {pdf_input} -> {out_path}")
        try:
            convert_pdf_to_excel(str(pdf_input), out_path)
        except Exception as e:
            print(f"Error during conversion: {e}")
            traceback.print_exc()
    else:
        print("[!] Error: No PDF input found.")
